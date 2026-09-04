#!/usr/bin/env python3
"""Regenerate availability_data.js from the latest weekly xlsx.

Sources:
  - Latest weekly xlsx (e.g. nativesonsexcelavail81726.xlsx) — plant list, sizes, prices, bloom/bud/new flags
  - Existing availability_data.js — full plant metadata (common name, height, origin, exposure, etc.)

Output:
  - availability_data.js — merged plant list with this week's availability

Logic:
  - For each plant in new xlsx:
    * If exists in current data: update sizes, bloom/bud/new flags; preserve all metadata
    * If not: add as new entry with minimal info (just sizes + flags)
  - Drop plants not in new xlsx (sold out this week)
  - Update week header and generated date
"""
from __future__ import annotations

import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent
XLSX = Path(sys.argv[1]) if len(sys.argv) > 1 else Path('/Users/tfross/.hermes/attachments/nativesonsexcelavail81726.xlsx')
AVAIL = ROOT / 'availability_data.js'
WEEK_HEADER = 'Week of September 7th, 2026'
GENERATED = '2026-09-04'


def norm(s: str) -> str:
    """Normalize plant name for fuzzy matching — also normalize for output consistency."""
    if s is None:
        return ''
    text = unicodedata.normalize('NFKC', str(s)).strip().lower()
    for old, new in (
        ('®', ''), ('™', ''), ('©', ''),
        ('\u2018', "'"), ('\u2019', "'"),
        ('\u201c', '"'), ('\u201d', '"'), ('\u00d7', ' x '),
    ):
        text = text.replace(old, new)
    return re.sub(r'\s+', ' ', text).strip()


def normalize_quotes(s: str) -> str:
    """Normalize curly quotes to straight quotes for output consistency with master DB."""
    if not s:
        return s
    text = str(s)
    for old, new in (
        ('\u2018', "'"), ('\u2019', "'"),
        ('\u201c', '"'), ('\u201d', '"'),
    ):
        text = text.replace(old, new)
    return text


def strip_pn(name: str) -> str:
    """Strip trailing P (patent) / N (native) markers used in the 4\" sheet."""
    return re.sub(r'\s+[PN]$', '', name).strip()


def parse_four_inch(ws):
    """Return dict[normalized_botanical] -> {botanical_orig, sizes, bloom, bud, new, section} from the 4\" sheet."""
    result = {}
    current_section = None
    current_price = None
    current_genus = None

    for r in range(4, ws.max_row + 1):
        val = ws.cell(r, 2).value
        if not val:
            continue
        s = str(val).strip()
        if not s:
            continue

        # Category header: "perennials   $3.15"
        if '$' in s:
            m = re.search(r'^(\w+)\s+\$([\d.]+)$', s)
            if m:
                current_section = m.group(1)
                current_price = float(m.group(2))
                current_genus = None
            continue

        # Strip ❀ bloom marker (keep P and N — they're patent/native codes)
        bloom = '❀' in s
        s = s.replace(' ❀', '').replace('❀', '').strip()
        if not s:
            continue

        first = s[0]
        # Standalone genus (e.g. "Ajuga") — establish genus context
        if first.isupper() and len(s.split()) == 1:
            current_genus = s
            continue
        # Cultivar-only (e.g. "'Chocolate Chip'") under current genus
        if first in ("'", '\u2018', '\u201c', '«'):
            full = f"{current_genus} {s}" if current_genus else s
        # Lowercase species (continuation under genus)
        elif first.islower():
            full = f"{current_genus} {s}" if current_genus else s
        # Full plant (e.g. "Geranium x cantabrigiense 'Biokovo'")
        else:
            parts = s.split(maxsplit=1)
            if parts and parts[0] not in ('x', 'X'):
                current_genus = parts[0]
            full = s

        # Strip trailing P/N markers and store original-cased name
        botanical_orig = strip_pn(full)

        if current_price is None:
            continue

        result[norm(botanical_orig)] = {
            'botanical': botanical_orig,
            'sizes': [{'container': '4in', 'price': current_price, 'order': ''}],
            'bloom': bloom,
            'bud': False,
            'new': False,
            'section': current_section or 'Perennial',
        }
    return result


def parse_main_sheet(ws):
    """Return dict[normalized_botanical] -> {botanical_orig, sizes, bloom, bud, new, section} from the 1g+ sheet."""
    result = {}
    for r in range(3, ws.max_row + 1):
        botanical = ws.cell(r, 1).value
        if not botanical:
            continue
        botanical = str(botanical).strip()
        if not botanical:
            continue

        status = ws.cell(r, 2).value
        status = str(status or '').strip()

        sizes = []
        for col, label in [(4, '1gal'), (6, '2gal'), (8, '5gal'), (10, '15gal')]:
            price = ws.cell(r, col).value
            if price and not str(price).startswith('*'):
                try:
                    sizes.append({'container': label, 'price': float(price), 'order': ''})
                except (ValueError, TypeError):
                    pass

        if not sizes:
            continue

        result[norm(botanical)] = {
            'botanical': botanical,
            'sizes': sizes,
            'bloom': ('Bloom' in status or 'Blm' in status),
            'bud': ('Bud' in status),
            'new': ('New' in status),
            'section': 'General Nursery',
        }
    return result


def main():
    if not XLSX.exists():
        raise SystemExit(f'Missing xlsx: {XLSX}')

    # Load existing data
    text = AVAIL.read_text(encoding='utf-8')
    m = re.search(r'window\.AVAILABILITY\s*=\s*(\{[\s\S]*?\});', text)
    if not m:
        raise SystemExit(f'Could not parse AVAILABILITY from {AVAIL}')
    # backfill_missing_enrichment.py injects fields via text edits and its
    # emitted lines used to end with a trailing comma (JS-style), which strict
    # json.loads rejects. Strip `,<newline><ws>}` / `,<newline><ws>]` before
    # decoding so both emitter styles round-trip. (Round 74, 2026-08-28.)
    payload = re.sub(r',(\s*\n\s*[}\]])', r'\1', m.group(1))
    data = json.loads(payload)

    # Build normalized lookup of existing plants
    existing = {}
    for p in data['plants']:
        key = norm(p['botanical'])
        existing.setdefault(key, []).append(p)

    # Parse new xlsx
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    new_4in = parse_four_inch(wb['4" material'])
    new_1g = parse_main_sheet(wb['1g and larger'])
    wb.close()

    # Normalize curly quotes in the new xlsx botanical names to straight quotes
    # so they match the master DB format and prior week's data style.
    for d in (new_4in, new_1g):
        for key in list(d.keys()):
            d[key]['botanical'] = normalize_quotes(d[key]['botanical'])

    # Merge: 1g+ wins over 4" (1g+ plants have richer pricing)
    new_data = {**new_4in, **new_1g}

    # Build merged plants list
    merged_plants = []
    matched_existing = set()
    new_added = 0
    for new_key, new_info in new_data.items():
        candidates = existing.get(new_key, [])
        if candidates:
            # Update existing plant
            plant = candidates[0]
            # Use the normalized-cased name from the new xlsx so existing data
            # adopts the same quote style as the source of truth.
            plant['botanical'] = new_info['botanical']
            plant['sizes'] = new_info['sizes']
            plant['bloom'] = new_info['bloom']
            plant['bud'] = new_info['bud']
            plant['new'] = new_info['new']
            merged_plants.append(plant)
            matched_existing.add(id(plant))
        else:
            # New plant not in existing — add with minimal info using original-cased name from xlsx
            merged_plants.append({
                'botanical': new_info['botanical'],
                'section': new_info.get('section', 'General Nursery'),
                'sizes': new_info['sizes'],
                'bloom': new_info['bloom'],
                'bud': new_info['bud'],
                'new': new_info['new'],
            })
            new_added += 1

    # Sort alphabetically by botanical name (stable ordering)
    merged_plants.sort(key=lambda p: p['botanical'].lower())

    # Update header
    data['week'] = WEEK_HEADER
    data['generated'] = GENERATED
    data['plants'] = merged_plants

    # Re-emit
    plants_json = json.dumps(merged_plants, indent=2, ensure_ascii=False)
    js = (
        '/* Native Sons Weekly Availability - generated */\n'
        '/*global window */\n'
        'window.AVAILABILITY = {\n'
        f'  "week": {json.dumps(WEEK_HEADER)},\n'
        f'  "generated": {json.dumps(GENERATED)},\n'
        f'  "source": {json.dumps(data.get("source", "Native Sons Wholesale Nursery weekly availability list"))},\n'
        f'  "contact": {json.dumps(data.get("contact", {"email": "orders@nativeson.com", "phone": "805.481.5996"}))},\n'
        f'  "plants": {plants_json}\n'
        '};'
    )
    AVAIL.write_text(js, encoding='utf-8')

    # Summary
    bloom_count = sum(1 for p in merged_plants if p.get('bloom'))
    bud_count = sum(1 for p in merged_plants if p.get('bud'))
    new_flag_count = sum(1 for p in merged_plants if p.get('new'))
    in_stock = len(merged_plants)
    sold_out = len(existing) - sum(1 for plist in existing.values() for p in plist if id(p) in matched_existing)
    print(f'✓ Generated {AVAIL}')
    print(f'  Plants this week: {in_stock}')
    print(f'  In bloom:         {bloom_count}')
    print(f'  Budding:          {bud_count}')
    print(f'  New:              {new_flag_count}')
    print(f'  Sold out (dropped): {sold_out}')
    print(f'  File size:        {AVAIL.stat().st_size:,} bytes')

    # Backfill missing enrichment from master catalog
    # Round 73 (2026-08-21): ships in scripts/backfill_missing_enrichment.py so weekly
    # refreshes self-heal when the master gets new entries or weekly entries land before
    # master keys. The script renames weekly → master canonical + fills blank fields.
    try:
        import subprocess
        backfill = Path(__file__).parent / 'scripts' / 'backfill_missing_enrichment.py'
        if backfill.exists():
            r = subprocess.run(
                ['python3', str(backfill), str(AVAIL)],
                capture_output=True, text=True, timeout=120
            )
            if r.returncode == 0:
                # Show only the summary lines from backfill, not the warnings
                for line in r.stdout.splitlines():
                    if any(k in line for k in ('Renames applied', 'Fields filled', 'Unmatched', 'Wrote')):
                        print(f'  backfill: {line.strip()}')
            else:
                print(f'  backfill warning: {r.stderr.strip().splitlines()[-1] if r.stderr else "non-zero exit"}')
    except Exception as e:
        print(f'  backfill skipped: {e}')


if __name__ == '__main__':
    main()
