#!/usr/bin/env python3
"""Sync bloom/bud/new flags from the weekly availability xlsx.

Usage:
  python3 scripts/sync_bloom_flags.py path/to/weekly.xlsx

Reads the 4" sheet (❀ marker on the plant name) and the 1g-and-larger sheet
(Status column = Bloom / Bud / New! / Blm / Bud 5) and overwrites
availability_data.js's bloom/bud/new flags.

The Excel is the source of truth for this week. The old
bud_bloom_overrides.json is no longer used (deleted 2026-07-24; the
in-bloom plant list drifted out of sync with the weekly xlsx, leaving 275
plants marked bloom instead of the actual 68 the xlsx records).
"""
from __future__ import annotations

import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl


def norm(s: str) -> str:
    if s is None:
        return ""
    text = unicodedata.normalize("NFKC", str(s)).strip().lower()
    for old, new in (
        ("®", ""), ("™", ""), ("©", ""),
        ("\u2018", "'"), ("\u2019", "'"),
        ("\u201c", '"'), ("\u201d", '"'), ("\u00d7", " x "),
    ):
        text = text.replace(old, new)
    return re.sub(r"\s+", " ", text).strip()


def parse_four_inch(ws):
    """Return dict[normalized_botanical] -> {bloom, bud, new} from the 4" sheet."""
    bloom = set()
    genus_stack: list[str] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not any(c is not None and str(c).strip() for c in row):
            continue
        full = " ".join(str(c) for c in row if c is not None)
        name = ""
        for c in row:
            if c is None:
                continue
            s = str(c).strip()
            if not s:
                continue
            if re.match(r"^\w+\s+\$", s):
                genus_stack = []
                name = ""
                break
            name = s
            break
        if not name:
            continue
        # Strip the in-flower (❀) and patented (P) / native (N) markers so the
        # resulting botanical matches the parser in build_order_portal_availability.py.
        name = re.sub(r"\s*\u2740\s*", " ", name)
        name = re.sub(r"\s+[PN]\s*$", "", name).strip()
        has_bloom = "\u2740" in full
        if re.match(r"^[A-Z][a-z]+$", name) and not has_bloom:
            genus_stack = [name]
            botanical = name
        else:
            if genus_stack and (re.match(r"^[a-z\u2018']", name)):
                botanical = f"{genus_stack[-1]} {name}"
            else:
                botanical = name
                first = name.split()[0] if name.split() else ""
                if re.match(r"^[A-Z][a-z]+$", first):
                    genus_stack = [first]
        if has_bloom:
            bloom.add(norm(botanical))
    return {"bloom": bloom, "bud": set(), "new": set()}


def parse_main_sheet(ws):
    """Return dict[normalized_botanical] -> {bloom, bud, new} from the 1g+ sheet."""
    bloom, bud, new = set(), set(), set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip()
        if not name or name.lower() == "nan":
            continue
        if not any(row[c] not in (None, "", 0) for c in (3, 5, 7, 9)):
            continue
        status = str(row[1] or "").strip()
        key = norm(name)
        if "Bloom" in status or "Blm" in status:
            bloom.add(key)
        if "Bud" in status:
            bud.add(key)
        if "New" in status:
            new.add(key)
    return {"bloom": bloom, "bud": bud, "new": new}


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit("Usage: sync_bloom_flags.py <weekly.xlsx> [availability_data.js]")
    xlsx = Path(sys.argv[1])
    avail = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(__file__).parent.parent / "availability_data.js"

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    four_inch = parse_four_inch(wb['4" material'])
    main = parse_main_sheet(wb["1g and larger"])
    wb.close()

    bloom_set = four_inch["bloom"] | main["bloom"]
    bud_set = main["bud"]
    new_set = main["new"]

    text = avail.read_text(encoding="utf-8")
    match = re.search(r"window\.AVAILABILITY\s*=\s*(\{[\s\S]*\});\s*$", text)
    if not match:
        raise SystemExit(f"Could not parse AVAILABILITY from {avail}")
    data = json.loads(match.group(1))

    before = {
        "bloom": sum(1 for p in data["plants"] if p.get("bloom")),
        "bud": sum(1 for p in data["plants"] if p.get("bud")),
        "new": sum(1 for p in data["plants"] if p.get("new")),
    }
    for p in data["plants"]:
        key = norm(p["botanical"])
        p["bloom"] = key in bloom_set
        p["bud"] = key in bud_set
        p["new"] = key in new_set
    after = {
        "bloom": sum(1 for p in data["plants"] if p.get("bloom")),
        "bud": sum(1 for p in data["plants"] if p.get("bud")),
        "new": sum(1 for p in data["plants"] if p.get("new")),
    }

    new_text = re.sub(
        r"window\.AVAILABILITY\s*=\s*\{[\s\S]*\};\s*$",
        lambda m: "window.AVAILABILITY = "
        + json.dumps(data, indent=2, ensure_ascii=False).replace("</", "<\\/")
        + ";\n",
        text,
    )
    avail.write_text(new_text, encoding="utf-8")

    print(f"Before: bloom={before['bloom']} bud={before['bud']} new={before['new']}")
    print(f"After:  bloom={after['bloom']} bud={after['bud']} new={after['new']}")
    print(f"Excel source: bloom={len(bloom_set)} bud={len(bud_set)} new={len(new_set)}")
    print(f"Wrote {avail} ({avail.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
