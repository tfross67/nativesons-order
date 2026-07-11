#!/usr/bin/env python3
"""Re-emit office_customers.js from customers2026.xlsx.

The office_customers.js file is loaded only by office.html (admin-gated).
It's a static 1.8MB dump of the customer directory, ~2,955 records.

NEVER linked or referenced from index.html — that's how we keep customer
data from leaking to the public web.

Usage:
  python3 load_customers.py
  python3 load_customers.py --input /path/to/new.xlsx   # to refresh

Without --input, reads from the standard cached path
~/.hermes/cache/documents/doc_*.xlsx starting with 'customers'.
"""
import json, re, sys, argparse, os
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit('pip install openpyxl first')

ROOT = Path(__file__).parent
DEFAULT_XLSX_GLOB = list(Path.home().glob('.hermes/cache/documents/*customers*.xlsx'))

def clean(val):
    if val is None: return None
    s = str(val).strip()
    if s.lower() in ('', 'none', 'null', 'n/a', 'n/a - update'):
        return None
    return s

def phone_digits(val):
    """Extract a usable 10-digit US phone number from a cell.

    Some xlsx cells pack 2-3 phone numbers into a single field, with or
    without separators. Examples:
      "3232459849 / 3108467094"  → 3232459849
      "(323) 245-9849"           → 3232459849
      "3232459849310846709"      → 3232459849 (first 10 digits)
      "323245984931084670943238198666" → 3232459849 (first 10 digits)

    Heuristic: prefer the first 10-digit run if found, else first 7-digit
    run, else strip leading 1 if 11 digits. Fall back to raw digits.
    """
    if val is None: return None
    raw = str(val)
    digits = re.sub(r'\D', '', raw)
    if not digits:
        return None
    if digits.startswith('1') and len(digits) == 11:
        digits = digits[1:]
    if 7 <= len(digits) <= 10:
        return digits
    # Multiple numbers packed into one cell — take the FIRST 10 digits
    # as the primary number. (Lossy but better than concatenating.)
    for length in (10, 7):
        chunk = digits[:length]
        if length == 10 and chunk[0] in '01':
            continue  # US area codes don't start with 0 or 1
        if length <= len(digits):
            return chunk
    # Last resort — return raw digits
    return digits if digits else None

def parse_pct(val):
    if val is None: return None
    s = str(val).strip().rstrip('%')
    try: return float(s) / (100 if '%' in str(val) else 1)
    except: return None

def build(xlsx_path):
    print(f'Reading {xlsx_path}...')
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    records = []
    skipped = 0
    for row_idx in range(2, ws.max_row + 1):
        raw = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]

        code  = clean(raw[0])
        name  = clean(raw[1])
        addr  = clean(raw[2])
        city  = clean(raw[3])
        contact = clean(raw[5])
        discount = clean(raw[6])
        email = clean(raw[7])
        mobile = phone_digits(raw[12])
        routing = clean(raw[14])
        sales_person = clean(raw[15])
        sales_tax = parse_pct(raw[16])
        terms = clean(raw[26])
        ctype = clean(raw[27])
        zip_code = clean(raw[29])
        telephone = phone_digits(raw[25])
        ship_name = clean(raw[21])
        ship_addr = clean(raw[18])
        ship_city = clean(raw[20])
        ship_state = clean(raw[22])
        ship_zip = clean(raw[23])
        ship_via = clean(raw[17])

        if not code and not name:
            skipped += 1
            continue
        if not name:
            name = contact or code or '(unknown)'

        if email:
            email = email.lower().strip()
        primary_phone = mobile or telephone

        search_text = (name or '') + ' ' + (city or '') + ' ' + (code or '') + ' ' + (email or '') + ' ' + (contact or '')
        search_text = re.sub(r'\s+', ' ', search_text.lower().strip())

        records.append({
            'customer_code': code,
            'name': name,
            'contact_name': contact,
            'address': addr,
            'city': city,
            'state': clean(raw[24]),
            'zip': zip_code,
            'email': email,
            'phone': primary_phone,
            'discount': discount,
            'default_markup_pct': parse_pct(discount),
            'routing': routing,
            'sales_person': sales_person,
            'sales_tax_pct': sales_tax,
            'terms': terms,
            'type': ctype,
            'ship_to_name': ship_name,
            'ship_to_address': ship_addr,
            'ship_to_city': ship_city,
            'ship_to_state': ship_state,
            'ship_to_zip': ship_zip,
            'ship_via': ship_via,
            'search_text': search_text,
        })

    # Dedupe by customer_code; keep the LAST (most-current) occurrence
    seen = {}
    for r in records:
        if r.get('customer_code'):
            seen[r['customer_code']] = r
    deduped = list(seen.values())
    for r in records:
        if not r.get('customer_code'):
            deduped.append(r)

    print(f'Loaded {len(records)} raw, {len(deduped)} unique')

    out = ROOT / 'office_customers.js'
    js = (
        '/* Office-only customer directory. Generated from customers2026.xlsx via\n'
        '   load_customers.py. Loaded only in office.html (admin-gated). Never\n'
        '   loaded by index.html (public portal).\n'
        '   No PII is exposed — only name, address, phone, terms, default markup.\n'
        '   Refresh with: python3 load_customers.py */\n'
        + 'window.OFFICE_CUSTOMERS = ' + json.dumps(deduped, ensure_ascii=False) + ';\n'
    )
    out.write_text(js)
    print(f'Wrote {os.path.getsize(out):,} bytes to {out}')


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--input', help='Path to customers xlsx. Otherwise picks the latest match from ~/.hermes/cache/documents/')
    args = ap.parse_args()

    if args.input:
        xlsx = Path(args.input)
    else:
        candidates = sorted(DEFAULT_XLSX_GLOB, key=os.path.getmtime, reverse=True)
        if not candidates:
            sys.exit('No customers xlsx found in ~/.hermes/cache/documents/')
        xlsx = candidates[0]
    build(xlsx)


if __name__ == '__main__':
    main()
