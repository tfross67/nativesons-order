#!/usr/bin/env python3
"""Backfill missing enrichment fields in availability_data.js from the master catalog.

Usage: python3 scripts/backfill_missing_enrichment.py [path/to/availability_data.js]
Defaults to /Users/tfross/.hermes/nativesons-order/availability_data.js.

This script is the last step of every weekly regen_availability.py run. It also
runs standalone after Irene sends a fresh master catalog. Two passes:

1. **Rename** weekly entries to match master canonical names
   (e.g. "Armeria 'Dreameria Daydream'" → "Armeria pseud. 'Dreameria\u2122 Daydream'").
2. **Enrichment** for plants that match the master (by normalized botanical name
   or alias) but have empty enrichment fields. Fills common, origin, height, width,
   hardiness, exposure, flower_color, flower_time, foliage, water, soil,
   special_uses, plant_type, description. Preserves ALL existing nonblank fields.

P3 rule: never manufacture matches. Plants with no master entry are skipped
silently (left exactly as they were).

Edits the file text-level so the existing JS object-literal structure (multi-line
pretty-printed, 4-space field indentation) is preserved. Each added field line is
inserted just before the plant's closing `  },` brace.
"""
from __future__ import annotations
import json, re, sys, unicodedata, shutil
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent
DEFAULT_AVAIL = ROOT / 'availability_data.js'
MASTER_DB = Path('/Users/tfross/.hermes/cache/documents/doc_c28bbe4bd06c_plantdatabase.xlsx')

# Add new entries here ONLY when a weekly name doesn't match the master key but
# is the same plant — and the master has the canonical form for it.
RENAMES = {
    "Armeria 'Dreameria Daydream'":  "Armeria pseud. 'Dreameria\u2122 Daydream'",
    "Armeria 'Dreameria Dreamland'": "Armeria pseud. 'Dreameria\u2122 Dreamland'",
}

# Aliases: weekly normalized key → master normalized key. Same key set as
# scripts/../../skills/.../enrich_portal_from_master.py so behaviour stays in
# sync with Tim's prior enrichment pass.
ALIASES = {
    "armeria p. 'dreameria dreamland'": "armeria pseud. 'dreameriatm dreamland'",
    "phyla nodiflora (pink)": "phyla nodiflora",
    "phyla nodiflora (white)": "phyla nodiflora 'white'",
    "nepeta faassenii 'whispurr pink'": "nepeta faassenii 'whispurrtm pink'",
    "campanula portenschlagiana 'resholt's variety'":
        "campanula portenschlagiana 'resholdt's variety'",
}

FIELD_MAP = [
    ('common', 'Common Name'),
    ('origin', 'Origin'),
    ('height', 'Height'),
    ('width', 'Width'),
    ('hardiness', 'Hardiness'),
    ('exposure', 'Exposure'),
    ('flower_color', 'Flower Color'),
    ('flower_time', 'Flower Time'),
    ('foliage', 'Foliage'),
    ('water', 'Water'),
    ('soil', 'Soil'),
    ('special_uses', 'Special Uses'),
    ('plant_type', 'Plant Type'),
    ('description', 'Additional Information'),
]


def norm(s):
    if s is None:
        return ''
    text = unicodedata.normalize('NFKC', str(s)).strip().lower()
    for old, new in (
        ('\u00ae', ''), ('\u2122', ''), ('\u00a9', ''),
        ('\u2018', "'"), ('\u2019', "'"),
        ('\u201c', '"'), ('\u201d', '"'), ('\u00d7', ' x '),
    ):
        text = text.replace(old, new)
    return re.sub(r'\s+', ' ', text).strip()


def nonblank(v):
    return v is not None and (not isinstance(v, str) or bool(v.strip()))


def clean(v):
    if not nonblank(v):
        return None
    return v.strip() if isinstance(v, str) else v


def prime_master_xlsx():
    """If canonical master path is missing, prime from the latest user upload (P46/P66)."""
    if MASTER_DB.exists():
        return
    cache = Path('/Users/tfross/.hermes/cache/documents')
    candidates = sorted(cache.glob('doc_*_plantdataexport.xlsx'),
                        key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidates:
        candidates = sorted(cache.glob('doc_*_plantmaterialsync.xlsx'),
                            key=lambda p: p.stat().st_mtime, reverse=True)
    if candidates:
        MASTER_DB.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy(candidates[0], MASTER_DB)
        print(f"Primed master xlsx from {candidates[0]}")


def load_master():
    prime_master_xlsx()
    wb = load_workbook(MASTER_DB, read_only=True, data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]
    master = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        name = row[2]
        if not name:
            continue
        padded = list(row) + [None] * (17 - len(row))
        rec = dict(zip(header, padded[:len(header)]))
        key = norm(name)
        richness = sum(1 for v in rec.values() if v)
        if key and (key not in master or richness > master[key].get('_r', 0)):
            rec['_r'] = richness
            master[key] = rec
    return master


def load_plants_via_node(path):
    """Load plants from the JS file via Node eval (handles object-literal syntax)."""
    import subprocess, tempfile
    helper = (
        'const fs = require("fs");\n'
        'const text = fs.readFileSync(process.argv[2], "utf8");\n'
        'const m = text.match(/window\\.AVAILABILITY\\s*=\\s*(\\{[\\s\\S]*?\\});\\s*$/m);\n'
        'if (!m) process.exit(1);\n'
        'const expr = m[1].replace(/;\\s*$/, "");\n'
        'eval("var avail = " + expr);\n'
        'process.stdout.write(JSON.stringify(avail.plants));\n'
    )
    with tempfile.NamedTemporaryFile(mode='w', suffix='.js', delete=False, dir='/tmp') as f:
        f.write(helper)
        helper_path = f.name
    try:
        result = subprocess.run(
            ['node', helper_path, str(path)],
            capture_output=True, text=True, timeout=30,
        )
        if result.returncode != 0:
            raise SystemExit(f"Node parse failed: {result.stderr}")
        return json.loads(result.stdout)
    finally:
        try:
            os.unlink(helper_path)
        except Exception:
            pass


def main():
    import argparse, os
    p = argparse.ArgumentParser()
    p.add_argument('path', nargs='?', default=str(DEFAULT_AVAIL))
    args = p.parse_args()
    path = Path(args.path)

    master = load_master()
    print(f"Master lookup: {len(master)} keys")

    plants = load_plants_via_node(path)
    print(f"Loaded {len(plants)} plants")

    # Plan renames + per-plant injection
    renames_applied = 0
    injections = {}  # botanical → list[(field_name, json_value_string)]
    unmatched = []

    for plant in plants:
        if plant['botanical'] in RENAMES:
            new = RENAMES[plant['botanical']]
            injections.setdefault('__renames__', []).append(
                (plant['botanical'], new)
            )
            plant['botanical'] = new
            renames_applied += 1

        key = norm(plant['botanical'])
        mk = ALIASES.get(key, key)
        rec = master.get(mk)
        if rec is None:
            unmatched.append(plant['botanical'])
            continue
        per_plant = []
        for pf, mf in FIELD_MAP:
            if nonblank(plant.get(pf)):
                continue
            v = clean(rec.get(mf))
            if v is None:
                continue
            if pf == 'hardiness' and isinstance(v, (int, float)):
                jv = json.dumps(str(v), ensure_ascii=False)
            else:
                jv = json.dumps(v, ensure_ascii=False)
            per_plant.append((pf, jv))
        if per_plant:
            injections[plant['botanical']] = per_plant

    # Apply text-level edits to the file
    text = path.read_text(encoding='utf-8')

    # 1) Renames (simple line replacements)
    rename_pairs = injections.pop('__renames__', [])
    for old, new in rename_pairs:
        old_line = f'    "botanical": "{old}",'
        new_line = f'    "botanical": "{new}",'
        if old_line in text:
            text = text.replace(old_line, new_line)

    # 2) Enrichment (per-plant injection)
    errors = []
    total_field_lines = 0
    for botanical, fields in injections.items():
        needle = f'    "botanical": "{botanical}",'
        b_idx = text.find(needle)
        if b_idx < 0:
            errors.append(f'no match for {botanical!r}')
            continue

        m_close = re.search(r'\n  \}(,?)\n', text[b_idx:])
        if not m_close:
            errors.append(f'no close brace for {botanical!r}')
            continue
        close_abs = b_idx + m_close.start()

        # Ensure the previous line ends with a comma (since we're adding fields after it)
        line_start = text.rfind('\n', 0, close_abs - 1) + 1
        prev_line = text[line_start:close_abs].rstrip()
        if not prev_line.endswith(',') and not prev_line.endswith('}') and not prev_line.endswith(']'):
            prev_line += ','
            text = text[:line_start] + prev_line + text[close_abs:]
            close_abs = line_start + len(prev_line)

        # Insert the new field lines right after the \n at close_abs.
        # Comma-separate the fields; the LAST field must NOT carry a trailing
        # comma or the file becomes JS-style (invalid for strict json.loads in
        # regen_availability.py / enrich_availability.py). Fixed round 74.
        inject_text = ',\n'.join(
            f'    "{fname}": {fval}' for fname, fval in fields
        ) + '\n'
        text = text[:close_abs + 1] + inject_text + text[close_abs + 1:]
        total_field_lines += len(fields)

    path.write_text(text, encoding='utf-8')

    print(f"Renames applied: {renames_applied}")
    print(f"Plants enriched: {len(injections) - sum(1 for _ in injections)}")  # minus the rename list
    print(f"Field lines added: {total_field_lines}")
    print(f"Unmatched (no master record): {len(unmatched)}")
    for b in unmatched:
        print(f"  {b}")
    if errors:
        print(f"Errors: {len(errors)}")
        for e in errors:
            print(f"  {e}")
    print(f"Wrote {path} ({path.stat().st_size:,} bytes)")


if __name__ == '__main__':
    main()
