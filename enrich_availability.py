#!/usr/bin/env python3
"""Enrich availability_data.js with metadata from plantdatabase.xlsx.

Reads the weekly availability parsed (already-dict-form) and the master
plant database for common name, origin, height, width, flower_color,
exposure, plant_type, water, soil, hardiness, special_uses, additional_info.

Usage:
  python3 enrich_availability.py [--fresh]

Reads from /tmp/merged_*.json (output of weekly parse step). If --fresh
is passed, parses xlsx and rebuilds. Otherwise enriches an existing
availability_data.js.

Match logic uses several variants of each botanical name:
  1. Normalized: lowercase, curly quotes → straight
  2. Strip parentheses content: "Phyla nodiflora (white)" → "phyla nodiflora"
  3. Strip quotes on cultivar
  4. Drop species epithet: "Thymus doerfleri 'Doone Valley'" → "thymus 'doone valley'"
"""
import json, re, os, sys, argparse
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent
MASTER_DB = Path('/Users/tfross/.hermes/cache/documents/doc_5465e03905a5_plantdatabase.xlsx')
AVAIL = ROOT / 'availability_data.js'


def variants(name):
    """Generate plausible key variants for matching."""
    if not name: return []
    s = str(name).strip()
    out = []
    # Normalized
    norm = s.lower().replace('\u2018', "'").replace('\u2019', "'")
    out.append(norm)
    # Strip parens content
    out.append(re.sub(r'\s*\([^)]*\)\s*', '', norm))
    # Strip quotes on cultivar
    out.append(re.sub(r"[\u2018\u2019\']", '', norm))
    # Drop species epithet
    m = re.match(r"^([A-Z][a-z]+)\s+[a-z\.]+(?:\s+x\s+[a-z]+)?\s+(.*)$", s)
    if m:
        out.append(f"{m.group(1).lower()} {m.group(2).lower()}")
    # Drop cultivar (keep binominal)
    m2 = re.match(r"^([A-Z][a-z]+\s+[a-z\.]+(?:\s+x\s+[a-z]+))\s+[\'\u2018\u2019]([^\'\u2018\u2019]+)[\'\u2018\u2019]$", s)
    if m2:
        out.append(m2.group(1).lower())
    return list(set(out))


def load_master():
    """Load master plant database to lookup keyed by normalized name."""
    if not MASTER_DB.exists():
        sys.exit(f'Missing {MASTER_DB}')
    wb = openpyxl.load_workbook(MASTER_DB, data_only=True)
    ws = wb.active
    lookup = {}
    for row_idx in range(2, ws.max_row + 1):
        row = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
        if not row[2]: continue
        record = {
            'common':           str(row[3]).strip()  if row[3]  else None,
            'origin':           str(row[4]).strip()  if row[4]  else None,
            'plant_type':       str(row[5]).strip()  if row[5]  else None,
            'exposure':         str(row[6]).strip()  if row[6]  else None,
            'flower_color':     str(row[7]).strip()  if row[7]  else None,
            'flower_time':      str(row[8]).strip()  if row[8]  else None,
            'height':           str(row[9]).strip()  if row[9]  else None,
            'width':            str(row[10]).strip() if row[10] else None,
            'foliage':          str(row[11]).strip() if row[11] else None,
            'water':            str(row[12]).strip() if row[12] else None,
            'hardiness':        row[13]               if row[13] else None,
            'soil':             str(row[14]).strip() if row[14] else None,
            'special_uses':     str(row[15]).strip() if row[15] else None,
            'additional_info':  str(row[16]).strip() if row[16] else None,
        }
        for v in variants(row[2]):
            lookup.setdefault(v, record)
    print(f'Loaded {len(lookup)} unique master records')
    return lookup


def enrich(plants, master):
    metadata_fields = [
        'common', 'origin', 'plant_type', 'exposure', 'flower_color', 'flower_time',
        'height', 'width', 'foliage', 'water', 'soil', 'special_uses', 'additional_info',
    ]
    enriched = 0
    for p in plants:
        matched = None
        for v in variants(p['botanical']):
            if v in master:
                matched = master[v]
                break
        if matched:
            for k in metadata_fields:
                val = matched.get(k)
                if val and not p.get(k):
                    p[k] = val
            h = matched.get('hardiness')
            if h is not None and not p.get('hardiness'):
                p['hardiness'] = f'{h}\u00b0F'
            enriched += 1
    return enriched


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--input', help='Path to JSON availability dict (default: existing availability_data.js)')
    args = ap.parse_args()

    master = load_master()

    if args.input:
        with open(args.input) as f:
            data = json.load(f)
    else:
        if not AVAIL.exists():
            sys.exit(f'No {AVAIL}')
        text = AVAIL.read_text()
        m = re.search(r'window\.AVAILABILITY = (\{[\s\S]*?\});', text)
        if not m:
            sys.exit(f'Could not parse AVAILABILITY from {AVAIL}')
        data = json.loads(m.group(1))

    n = enrich(data['plants'], master)
    print(f'Enriched {n}/{len(data["plants"])} plants')

    # Re-emit
    plants_json = json.dumps(data['plants'], indent=2, ensure_ascii=False)
    js = (
        '/* Native Sons Weekly Availability - generated */\n'
        '/*global window */\n'
        'window.AVAILABILITY = {\n'
        f'  "week": {json.dumps(data["week"])},\n'
        f'  "generated": {json.dumps(data["generated"])},\n'
        f'  "source": {json.dumps(data["source"])},\n'
        f'  "contact": {json.dumps(data["contact"])},\n'
        f'  "plants": {plants_json}\n'
        '};\n'
    )
    AVAIL.write_text(js)
    print(f'Wrote {os.path.getsize(AVAIL)} bytes to {AVAIL}')


if __name__ == '__main__':
    main()
